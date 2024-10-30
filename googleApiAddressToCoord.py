import requests
import openpyxl
from config import GOOGLE_MAPS_API_KEY
import os
import logging
from datetime import datetime

def setup_logging():
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"geocoding_log_{timestamp}.txt")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger()

def validate_api_key(api_key):
    try:
        # Attempt a simple geocode to validate the key
        url = f'https://maps.googleapis.com/maps/api/geocode/json?address=1600+Amphitheatre+Parkway,+Mountain+View,+CA&key={api_key}'
        response = requests.get(url)
        data = response.json()
        if data['status'] == 'OK':
            return True
        else:
            logger.error(f"API Key validation failed: {data['status']}")
            logger.error("Please check your API key and ensure it's correctly set up in the Google Cloud Console.")
            logger.error("1. Visit https://console.cloud.google.com/")
            logger.error("2. Select your project")
            logger.error("3. Go to 'APIs & Services' > 'Credentials'")
            logger.error("4. Check if the API key is correctly created and has the necessary permissions")
            logger.error("5. Ensure the Geocoding API is enabled for your project")
            logger.error("6. If you've recently created the key, wait a few minutes for it to activate")
            return False
    except Exception as e:
        logger.error(f"An unexpected error occurred while validating the API key: {str(e)}")
        return False

def geocode_address(address):
    """
    Retrieves the latitude and longitude coordinates for a given address using the Google Maps Geocoding API.
    
    Args:
        address (str): The address to be converted to coordinates.
    
    Returns:
        tuple: A tuple containing the latitude and longitude coordinates, or (None, None) if the geocoding fails.
    """
    url = f'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={GOOGLE_MAPS_API_KEY}'
    try:
        response = requests.get(url)
        data = response.json()
        if data['status'] == 'OK':
            location = data['results'][0]['geometry']['location']
            return (location['lat'], location['lng'])
        else:
            logger.warning(f"No results found for address: {address}")
            return (None, None)
    except Exception as e:
        logger.error(f"Error geocoding address: {address}. Error: {str(e)}")
        return (None, None)

def main():
    global logger
    logger = setup_logging()

    # Validate the Google Maps API key
    if not validate_api_key(GOOGLE_MAPS_API_KEY):
        logger.error("Exiting due to invalid API key.")
        return

    # Load addresses from Excel file
    workbook = openpyxl.load_workbook('addresses.xlsx')
    sheet = workbook.active

    # Create a new workbook for output
    wb_out = openpyxl.Workbook()
    sheet_out = wb_out.active
    sheet_out.append(['Address', 'Latitude', 'Longitude', 'Status'])

    total_addresses = 0
    successful_geocodes = 0
    failed_geocodes = 0

    # Assume the first row is a header and start from the second row
    for row in range(2, sheet.max_row + 1):
        address = sheet.cell(row=row, column=8).value
        if address:
            total_addresses += 1
            logger.info(f"Geocoding: {address}")
            latitude, longitude = geocode_address(address)
            if latitude is not None and longitude is not None:
                sheet_out.append([address, latitude, longitude, 'Success'])
                successful_geocodes += 1
            else:
                sheet_out.append([address, 'Not found', 'Not found', 'Failed'])
                failed_geocodes += 1

    # Save the updated workbook to a new Excel file
    output_filename = 'address_coordinates.xlsx'
    try:
        wb_out.save(output_filename)
        logger.info(f"Geocoding complete. Results written to '{output_filename}'")
        logger.info(f"Total addresses processed: {total_addresses}")
        logger.info(f"Successful geocodes: {successful_geocodes}")
        logger.info(f"Failed geocodes: {failed_geocodes}")
    except Exception as e:
        logger.error(f"Error saving output file: {e}")

if __name__ == "__main__":
    main()
